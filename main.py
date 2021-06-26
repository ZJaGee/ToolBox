from openpyxl import Workbook, load_workbook
import os
from tempfile import NamedTemporaryFile
from wxpy import *
from wechaty import Wechaty
from moviepy.editor import *
import asyncio
from PyQt5.QtWidgets import QApplication, QWidget


def main():
    templateDir = os.path.dirname(os.path.abspath(__file__))

    wb = Workbook()
    # sheet = wb[wb.sheetnames[0]]
    # sheet['A1'] = 'test'
    shDetail = wb.create_sheet()
    shDetail.title = "new"
    shDetail['A1'] = 'test'
    wb.save(os.path.join(templateDir, "tmp.xlsx"))

    # wb = load_workbook(os.path.join(templateDir, "tmp.xlsx"))
    # sheet = wb[wb.sheetnames[0]]
    # sheet['A1'] = 'Hello world!'
    # shDetail = wb.create_sheet()
    # shDetail.title = "new"
    # wb.save(os.path.join(templateDir, "tmp.xlsx"))
    # xlsxMd5Str, fileStream = "", bytes([])
    with NamedTemporaryFile() as tmp:
        tmp.close()
        print(tmp.name)
        wb.save(tmp.name)
        # tmp.seek(0)
        # fileStream = tmp.read()
        # xlsxMd5Str = md5bytes(fileStream)

    # return xlsxMd5Str, len(fileStream), fileStream


# def controlPhone():
class MyBot(Wechaty):
    async def on_message(self, msg: Message):
        talker = msg.talker()
        await talker.ready()
        talker.say('测试')


async def main():
    bot = MyBot()
    await bot.start()


async def ChatTest():
    bot = Wechaty()
    bot.on('scan', lambda status, qrcode, data: print(
        'Scan QR Code to login: {}\nhttps://wechaty.js.org/qrcode/{}'.format(status, qrcode)))
    bot.on('login', lambda user: print('User {} logged in'.format(user)))
    bot.on('message', lambda message: print('Message: {}'.format(message)))
    await bot.start()


def receiveMusic(musicName, filePath):
    musicName = musicName.split('.mp4')[0]
    print(musicName)
    # video = VideoFileClip("../src/src/"+musicName+'.mp4')
    video = VideoFileClip(filePath)
    audio = video.audio
    audio.write_audiofile("../src/dir/wav/"+musicName+'.wav')
    audio.write_audiofile("../src/dir/mp3/"+musicName+'.mp3')


if __name__ == '__main__':
    # main()
    # controlPhone()
    # bot = Bot()
    # my_friend = bot.friends().search('郑佳煜', sex=MALE, city="广州")[0]
    # my_friend.send('测试')
    # asyncio.run(ChatTest())
    # receiveMusic()
    # path = "../src/src"
    # for root, dirs, files in os.walk(path):
    #     print("files:", files)
    #     for file in files:
    #         print(os.path.join(root, file))
    #         filePath = os.path.join(root, file)
    #         receiveMusic(file, filePath)

    app = QApplication(sys.argv)
    # QWidget部件是pyqt5所有用户界面对象的基类。他为QWidget提供默认构造函数。默认构造函数没有父类。
    w = QWidget()
    # resize()方法调整窗口的大小。这离是250px宽150px高
    w.resize(250, 150)
    # move()方法移动窗口在屏幕上的位置到x = 300，y = 300坐标。
    w.move(300, 300)
    # 设置窗口的标题
    w.setWindowTitle('Simple')
    # 显示在屏幕上
    w.show()

    # 系统exit()方法确保应用程序干净的退出
    # 的exec_()方法有下划线。因为执行是一个Python关键词。因此，exec_()代替
    sys.exit(app.exec_())
